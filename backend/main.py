import os
import threading
import traceback
import shutil
import time
from io import BytesIO
from typing import List, Optional
import openpyxl
from fastapi import FastAPI, Request, BackgroundTasks, UploadFile, File, Form, Depends, HTTPException
from fastapi.templating import Jinja2Templates
from fastapi.responses import JSONResponse, RedirectResponse, FileResponse, StreamingResponse
from fastapi import Response, status
from dotenv import load_dotenv
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import select, text
from db import engine, Base, get_db
from security import hash_password, verify_password, create_access_token
from models import User, QuestionPaper
from schemas import UserCreate, UserLogin
from dependencies import get_current_user
from ocr import extract_handwritten_text, extract_answers_by_question
from vector_store import upsert_model_answers, has_embeddings, get_similarity, delete_qp_vectors, get_embedding_status_bulk
from gemini_service import generate_model_answers, extract_marks
from vector_store import fetch_model_answers as _fetch_model_answers

MAX_QUESTIONS = 4       # student must attempt 4 out of 6
MARKS_PER_QUESTION = 5  # each question is worth 5 marks
MAX_TOTAL_MARKS = MAX_QUESTIONS * MARKS_PER_QUESTION  # 20


def _round_half(x: float) -> float:
    """Round to nearest 0.5 (e.g. 3.3 → 3.5, 3.7 → 4.0, 4.8 → 5.0)."""
    return round(x * 2) / 2


app = FastAPI()

progress_tracker: dict = {}
_progress_lock = threading.Lock()

@app.get("/progress/question-paper/{filename}")
def get_progress(filename: str):
    return progress_tracker.get(filename, {"status": "Not started", "progress": 0})

def _set_progress(filename: str, status: str, progress: int):
    with _progress_lock:
        progress_tracker[filename] = {"status": status, "progress": progress}

def process_question_paper_task(filename: str, file_path: str, qp_id: int):
    try:
        _set_progress(filename, "Extracting text from document...", 10)

        # 1. Extract text — already returns a structured {Q1: text, ...} dict
        questions = extract_handwritten_text(file_path)

        _set_progress(filename, "Generating model answers...", 60)

        # 2. Generate answers and extract marks
        model_answers = generate_model_answers(questions)
        marks_map = {q_num: extract_marks(q_text) for q_num, q_text in questions.items()}

        _set_progress(filename, "Saving embeddings to Vector DB...", 90)

        # 3. Upsert vectors
        upsert_model_answers(qp_id, model_answers, questions, marks_map)

        _set_progress(filename, "Completed successfully!", 100)
    except Exception as e:
        traceback.print_exc()
        _set_progress(filename, f"Error: {str(e)}", -1)

# pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

templates = Jinja2Templates(directory="templates")

BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(BACKEND_DIR)

env_path = os.path.join(BASE_DIR, ".env")

load_dotenv(env_path)

QUES_PATH = "question_papers"
ANS_PATH = "answer_sheets"

@app.on_event("startup")
def startup():
    Base.metadata.create_all(bind=engine)
    # Add qp_display_name column to existing tables that predate this migration
    with engine.connect() as conn:
        try:
            conn.execute(text("ALTER TABLE question_papers ADD COLUMN qp_display_name VARCHAR(255) NULL"))
            conn.commit()
            print("Migration: added qp_display_name column")
        except Exception:
            pass  # Column already exists — safe to ignore

@app.get("/")
def read_root(request: Request):
    return templates.TemplateResponse("home.html", {"request": request})

@app.get("/register")
def register(request: Request):
    return templates.TemplateResponse("register.html", {"request": request})

@app.post("/register")
def register_user(
    user: UserCreate,
    db: Session = Depends(get_db)
):

    result = db.execute(
        select(User).where(User.email == user.email)
    )
    existing_user = result.scalar_one_or_none()

    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")

    new_user = User(
        username=user.username,
        email=user.email,
        hashed_password=hash_password(user.password)
    )

    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    return {"message": "User registered successfully"}

@app.get("/login")
async def login(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})

@app.post("/login")
def login_user(
    response: Response,
    user: UserLogin,
    db: Session = Depends(get_db)
):

    result = db.execute(
        select(User).where(User.email == user.email)
    )

    db_user = result.scalar_one_or_none()

    if not db_user or not verify_password(user.password, db_user.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    token = create_access_token({
        "sub": db_user.email,
        "username": db_user.username,
        "id": db_user.id
    })

    response.set_cookie(
        key="access_token",
        value=token,
        httponly=True,
        samesite="lax"
    )

    return {"message": "Login successful"}

@app.get("/logout")
async def logout(response: Response):
    response.delete_cookie("access_token")
    # Redirect to login page
    return RedirectResponse(url="/login", status_code=status.HTTP_302_FOUND)


@app.get("/dashboard")
async def dashboard(
    request: Request,
    user=Depends(get_current_user)
):
    return templates.TemplateResponse(
        "dashboard.html",
        {"request": request, "username": user["username"]} 
    )

@app.get("/evaluate")
async def evaluate(
    request: Request,
    user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # Fetch question papers for the logged-in user
    result = db.execute(
        select(QuestionPaper).where(QuestionPaper.user_id == user["id"])
    )
    question_papers = result.scalars().all()
    
    return templates.TemplateResponse(
        "evaluate.html",
        {"request": request, "question_papers": question_papers}
    )

@app.post("/upload-question-paper")
def upload_question_paper(
    background_tasks: BackgroundTasks,
    questionPaper: UploadFile = File(...),
    display_name: str = Form(...),
    user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    result = db.execute(
        select(QuestionPaper).where(
            QuestionPaper.qp_name == questionPaper.filename,
            QuestionPaper.user_id == user["id"]
        )
    )
    existing_qp = result.scalar_one_or_none()

    ext = os.path.splitext(questionPaper.filename)[1]
    file_bytes = questionPaper.file.read()

    if existing_qp:
        # Re-upload: overwrite the existing file at its stored path
        file_location = existing_qp.qp_path
        with open(file_location, "wb") as buffer:
            buffer.write(file_bytes)
        existing_qp.qp_display_name = display_name
        db.commit()
        _set_progress(questionPaper.filename, "Starting reprocessing...", 5)
        background_tasks.add_task(process_question_paper_task, questionPaper.filename, file_location, existing_qp.qp_id)
        return {"message": "Reprocessing started", "filename": questionPaper.filename}

    # New upload — use a unique filename to avoid cross-user collisions
    unique_filename = f"qp_{user['id']}_{int(time.time())}{ext}"
    file_location = f"{QUES_PATH}/{unique_filename}"
    with open(file_location, "wb") as buffer:
        buffer.write(file_bytes)

    new_qp = QuestionPaper(
        qp_name=questionPaper.filename,
        qp_display_name=display_name,
        qp_path=file_location,
        user_id=user["id"]
    )

    db.add(new_qp)
    db.commit()
    db.refresh(new_qp)

    _set_progress(questionPaper.filename, "Starting processing...", 5)
    background_tasks.add_task(process_question_paper_task, questionPaper.filename, file_location, new_qp.qp_id)

    return {"message": "Upload started", "filename": questionPaper.filename}

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.post("/evaluate")
async def evaluate(
    paper_id: int = Form(...),
    answerSheet: UploadFile = File(...),
    user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        print("Received paper_id:", paper_id)
        print("Uploaded file:", answerSheet.filename)

        # --- Fetch question paper from DB ---
        question_paper = db.query(QuestionPaper).filter(
            QuestionPaper.qp_id == paper_id
        ).first()

        if not question_paper:
            return JSONResponse(
                status_code=404,
                content={"detail": "Question paper not found"}
            )

        # --- Check embeddings exist ---
        if not has_embeddings(paper_id):
            return JSONResponse(
                status_code=400,
                content={"detail": "Question paper has not been processed yet. Please wait for processing to complete before evaluating."}
            )

        # --- Save uploaded answer sheet ---
        # Give unique name to avoid filename collisions
        ext = os.path.splitext(answerSheet.filename)[1]
        unique_filename = f"ans_{paper_id}_{int(time.time())}{ext}"
        answer_path = os.path.join(UPLOAD_FOLDER, unique_filename)

        with open(answer_path, "wb") as buffer:
            shutil.copyfileobj(answerSheet.file, buffer)

        print("Answer sheet saved to:", answer_path)

        # --- Extract answers grouped by question ---
        answers_by_question = extract_answers_by_question(answer_path)

        print("\n===== ANSWERS BY QUESTION =====\n", answers_by_question)

        # Pull out cover-page metadata (not question answers)
        METADATA_KEYS = {"roll_number", "exam_name", "subject", "class_name", "date_of_exam"}
        roll_number  = answers_by_question.get("roll_number", "")
        exam_name    = answers_by_question.get("exam_name", "")
        subject      = answers_by_question.get("subject", "")
        class_name   = answers_by_question.get("class_name", "")
        date_of_exam = answers_by_question.get("date_of_exam", "")

        # --- Evaluate each answered question using Pinecone vector similarity ---
        scored = []   # questions that were successfully graded
        errors = []   # questions that failed (no model answer found)

        for question_number, student_answer in answers_by_question.items():
            if question_number in METADATA_KEYS:
                continue
            if not student_answer or not isinstance(student_answer, str) or not student_answer.strip():
                print(f"Skipping {question_number}: empty answer from OCR")
                continue

            try:
                similarity_data = get_similarity(paper_id, question_number, student_answer)

                cosine_sim    = similarity_data["cosine_similarity"]
                awarded_marks = _round_half((cosine_sim**0.6) * MARKS_PER_QUESTION)

                scored.append({
                    "question_number": question_number,
                    "question_text": similarity_data["question_text"],
                    "student_answer": student_answer,
                    "model_answer": similarity_data["model_answer"],
                    "cosine_similarity_pct": similarity_data["cosine_similarity_pct"],
                    "max_marks": MARKS_PER_QUESTION,
                    "awarded_marks": awarded_marks
                })

            except Exception as e:
                print(f"Skipping {question_number}: {e}")
                errors.append({
                    "question_number": question_number,
                    "student_answer": student_answer,
                    "max_marks": MARKS_PER_QUESTION,
                    "awarded_marks": 0,
                    "error": "No model answer found for this question"
                })

        # --- Best-of-4 selection (6 questions, attempt any 4) ---
        # Sort by awarded_marks descending and keep the top MAX_QUESTIONS
        scored.sort(key=lambda r: r["awarded_marks"], reverse=True)
        selected = scored[:MAX_QUESTIONS]

        total_score = sum(r["awarded_marks"] for r in selected)
        results     = selected + errors   # show selected answers then any errors

        return {
            "total_score": _round_half(total_score),
            "max_marks": MAX_TOTAL_MARKS,
            "questions_counted": len(selected),
            "roll_number": roll_number,
            "exam_name": exam_name,
            "subject": subject,
            "class_name": class_name,
            "date_of_exam": date_of_exam,
            "results": results
        }

    except Exception as e:
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={"detail": str(e)}
        )



# ── My Papers ─────────────────────────────────────────────────────────────────

@app.get("/my-papers")
async def my_papers(
    request: Request,
    user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    result = db.execute(
        select(QuestionPaper).where(QuestionPaper.user_id == user["id"])
    )
    qps = result.scalars().all()

    # Single Pinecone call for all papers
    qp_ids = [qp.qp_id for qp in qps]
    status_map = get_embedding_status_bulk(qp_ids) if qp_ids else {}

    papers = [
        {
            "qp_id": qp.qp_id,
            "display_name": qp.qp_display_name or qp.qp_name,
            "qp_name": qp.qp_name,
            "created_at": qp.created_at,
            "vectors_ready": status_map.get(qp.qp_id, False),
        }
        for qp in qps
    ]

    return templates.TemplateResponse(
        "my_papers.html",
        {"request": request, "papers": papers, "username": user["username"]}
    )


# ── File serving ───────────────────────────────────────────────────────────────

def _get_owned_qp(qp_id: int, user_id: int, db: Session) -> QuestionPaper:
    result = db.execute(
        select(QuestionPaper).where(
            QuestionPaper.qp_id == qp_id,
            QuestionPaper.user_id == user_id
        )
    )
    qp = result.scalar_one_or_none()
    if not qp:
        raise HTTPException(status_code=404, detail="Question paper not found")
    abs_path = os.path.join(BACKEND_DIR, qp.qp_path)
    if not os.path.exists(abs_path):
        raise HTTPException(status_code=404, detail="File not found on disk")
    qp._abs_path = abs_path
    return qp


@app.get("/question-papers/view/{qp_id}")
async def view_question_paper(
    qp_id: int,
    user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    qp = _get_owned_qp(qp_id, user["id"], db)
    ext = os.path.splitext(qp.qp_path)[1].lower()
    mime = "application/pdf" if ext == ".pdf" else f"image/{ext.lstrip('.')}"
    return FileResponse(qp._abs_path, media_type=mime, headers={"Content-Disposition": "inline"})


@app.get("/question-papers/download/{qp_id}")
async def download_question_paper(
    qp_id: int,
    user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    qp = _get_owned_qp(qp_id, user["id"], db)
    filename = os.path.basename(qp.qp_path)
    return FileResponse(
        qp._abs_path,
        filename=filename,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── QP Deletion ────────────────────────────────────────────────────────────────

@app.delete("/question-papers/{qp_id}")
async def delete_question_paper(
    qp_id: int,
    user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        result = db.execute(
            select(QuestionPaper).where(
                QuestionPaper.qp_id == qp_id,
                QuestionPaper.user_id == user["id"]
            )
        )
        qp = result.scalar_one_or_none()
        if not qp:
            return JSONResponse(status_code=404, content={"detail": "Question paper not found"})

        abs_path = os.path.join(BACKEND_DIR, qp.qp_path)
        if os.path.exists(abs_path):
            os.remove(abs_path)

        try:
            delete_qp_vectors(qp_id)
        except Exception as e:
            print(f"Pinecone delete failed for QP {qp_id}: {e}")

        db.delete(qp)
        db.commit()

        return JSONResponse(status_code=200, content={"message": "Deleted successfully"})

    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"detail": str(e)})


# ── XLSX Report Generation ─────────────────────────────────────────────────────

class ReportEntry(BaseModel):
    roll_number: str = ""
    exam_name: str = ""
    class_name: str = ""
    subject: str = ""
    date_of_exam: str = ""
    total_score: float = 0.0
    max_marks: float = 0.0


@app.post("/generate-report")
async def generate_report(
    entries: List[ReportEntry],
    user=Depends(get_current_user)
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # Header row
    headers = ["H.T. No.", "Exam", "Branch", "Subject", "Date", "Score", "Max Marks", "Percentage"]
    ws.append(headers)

    # Style header
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="06B6D4", end_color="06B6D4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for e in entries:
        pct = round((e.total_score / e.max_marks * 100), 2) if e.max_marks > 0 else 0.0
        ws.append([e.roll_number, e.exam_name, e.class_name, e.subject, e.date_of_exam, e.total_score, e.max_marks, pct])

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    # File name based on first entry's class + subject
    if entries:
        safe = lambda s: "".join(c if c.isalnum() or c in "-_" else "_" for c in s)
        fname = f"{safe(entries[0].class_name)}_{safe(entries[0].subject)}.xlsx"
    else:
        fname = "report.xlsx"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )
